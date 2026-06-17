#!/usr/bin/env python3
"""
generate_report.py — Aggregate Vitest JSON + pytest XML into a single Excel report.

Usage:
    python tests/generate_report.py \
        --vitest  reports/vitest-results.json \
        --pytest  reports/pytest-results.xml  \
        --out     reports/test-report-<timestamp>.xlsx
"""
import argparse
import json
import os
import re
import sys
import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    print("[ERROR] openpyxl not installed — run: pip install openpyxl")
    sys.exit(1)

try:
    import xml.etree.ElementTree as ET
except ImportError:
    ET = None  # type: ignore

# ── Colours ──────────────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
BLUE   = PatternFill("solid", fgColor="BDD7EE")
GREY   = PatternFill("solid", fgColor="D9D9D9")

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E7D32")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ── Helpers ───────────────────────────────────────────────────────────────────
def _autowidth(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value or ""))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)


def _header_row(ws, headers: list[str]):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22


def _fill_cell(cell, fill: PatternFill):
    cell.fill = fill
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def _status_fill(status: str) -> PatternFill:
    s = status.lower()
    if s in ("pass", "passed"):
        return GREEN
    if s in ("fail", "failed"):
        return RED
    if s in ("skip", "skipped"):
        return YELLOW
    return GREY


# ── Vitest JSON parser ────────────────────────────────────────────────────────
def parse_vitest(path: str) -> list[dict]:
    """Parse Vitest JSON reporter output → list of test result dicts."""
    if not path or not os.path.exists(path):
        return []
    with open(path) as f:
        data = json.load(f)

    results = []
    test_results = data.get("testResults", [])
    for suite_file in test_results:
        suite_name = os.path.basename(suite_file.get("name", "unknown"))
        for t in suite_file.get("assertionResults", []):
            status_raw = t.get("status", "unknown")
            status = "PASSED" if status_raw == "passed" else ("FAILED" if status_raw == "failed" else "SKIPPED")
            error = ""
            if t.get("failureMessages"):
                error = "\n".join(t["failureMessages"])[:500]
            results.append({
                "suite": suite_name,
                "name": " > ".join(t.get("ancestorTitles", []) + [t.get("title", "")]),
                "status": status,
                "duration_ms": t.get("duration", 0) or 0,
                "error": error,
                "source": "component",
                "screenshot": "",
            })
    return results


# ── pytest XML parser ─────────────────────────────────────────────────────────
def parse_pytest(path: str) -> list[dict]:
    """Parse pytest JUnit XML → list of test result dicts."""
    if not path or not os.path.exists(path):
        return []
    tree = ET.parse(path)
    root = tree.getroot()
    results = []

    for testsuite in root.iter("testsuite"):
        suite_name = testsuite.get("name", "e2e")
        for tc in testsuite.iter("testcase"):
            name = f"{tc.get('classname', '')}.{tc.get('name', '')}".lstrip(".")
            duration_s = float(tc.get("time", 0) or 0)
            failure = tc.find("failure")
            skip    = tc.find("skipped")
            error   = tc.find("error")

            if failure is not None:
                status = "FAILED"
                err_msg = (failure.get("message") or "") + "\n" + (failure.text or "")
            elif skip is not None:
                status = "SKIPPED"
                err_msg = skip.get("message", "")
            elif error is not None:
                status = "FAILED"
                err_msg = (error.get("message") or "") + "\n" + (error.text or "")
            else:
                status = "PASSED"
                err_msg = ""

            # Extract screenshot path from stdout if present
            screenshot = ""
            for sys_out in tc.iter("system-out"):
                m = re.search(r"\[SCREENSHOT\]\s+(.+\.png)", sys_out.text or "")
                if m:
                    screenshot = m.group(1).strip()

            results.append({
                "suite": suite_name,
                "name": name,
                "status": status,
                "duration_ms": round(duration_s * 1000),
                "error": err_msg[:500],
                "source": "e2e",
                "screenshot": screenshot,
            })
    return results


# ── Sheet builders ─────────────────────────────────────────────────────────────
def build_summary(wb, all_tests: list[dict], duration_s: float, timestamp: str):
    ws = wb.active
    ws.title = "Summary"

    total   = len(all_tests)
    passed  = sum(1 for t in all_tests if t["status"] == "PASSED")
    failed  = sum(1 for t in all_tests if t["status"] == "FAILED")
    skipped = sum(1 for t in all_tests if t["status"] == "SKIPPED")
    rate    = f"{(passed / total * 100):.1f}%" if total else "N/A"

    rows = [
        ("Metric", "Value"),
        ("Run timestamp", timestamp),
        ("Total tests", total),
        ("Passed",  passed),
        ("Failed",  failed),
        ("Skipped", skipped),
        ("Pass rate", rate),
        ("Total duration (s)", f"{duration_s:.1f}"),
        ("", ""),
        ("Layer", "Tests"),
        ("Component (Vitest)", sum(1 for t in all_tests if t["source"] == "component")),
        ("E2E Selenium",       sum(1 for t in all_tests if t["source"] == "e2e")),
    ]

    for i, row in enumerate(rows, 1):
        ws.append(row)
        for cell in ws[i]:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    # Colour the value cells in summary rows
    STATUS_ROW = {"Passed": GREEN, "Failed": RED, "Skipped": YELLOW}
    for i, (label, _) in enumerate(rows, 1):
        if label in STATUS_ROW:
            ws.cell(i, 2).fill = STATUS_ROW[label]
        if label in ("Metric", "Layer"):
            ws.cell(i, 1).font = Font(bold=True)
            ws.cell(i, 1).fill = HEADER_FILL
            ws.cell(i, 1).font = HEADER_FONT
            ws.cell(i, 2).fill = HEADER_FILL
            ws.cell(i, 2).font = HEADER_FONT

    # Big pass rate cell
    rate_row = next((i + 1 for i, (l, _) in enumerate(rows) if l == "Pass rate"), None)
    if rate_row:
        cell = ws.cell(rate_row, 2)
        cell.font = Font(bold=True, size=14)
        cell.fill = GREEN if passed == total else (RED if failed > 0 else YELLOW)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20


def build_details(wb, all_tests: list[dict]):
    ws = wb.create_sheet("Details")
    _header_row(ws, ["#", "Source", "Suite", "Test Name", "Status", "Duration (ms)", "Error"])

    for i, t in enumerate(all_tests, 1):
        ws.append([
            i,
            t["source"].upper(),
            t["suite"],
            t["name"],
            t["status"],
            t["duration_ms"],
            t["error"],
        ])
        fill = _status_fill(t["status"])
        for col in range(1, 8):
            _fill_cell(ws.cell(i + 1, col), fill)

    _autowidth(ws)


def build_failures(wb, all_tests: list[dict]):
    ws = wb.create_sheet("Failures")
    failed = [t for t in all_tests if t["status"] == "FAILED"]

    if not failed:
        ws.append(["No failures — all tests passed! 🎉"])
        ws["A1"].font = Font(bold=True, size=14, color="2E7D32")
        return

    _header_row(ws, ["#", "Source", "Suite", "Test Name", "Error Message", "Screenshot"])
    for i, t in enumerate(failed, 1):
        ws.append([i, t["source"].upper(), t["suite"], t["name"], t["error"], t["screenshot"]])
        for col in range(1, 7):
            _fill_cell(ws.cell(i + 1, col), RED)

    _autowidth(ws)


def build_by_layer(wb, all_tests: list[dict]):
    ws = wb.create_sheet("By Layer")
    _header_row(ws, ["Layer", "Total", "Passed", "Failed", "Skipped", "Pass Rate"])

    layers = {"component": "Component (Vitest)", "e2e": "E2E (Selenium)"}
    for key, label in layers.items():
        subset = [t for t in all_tests if t["source"] == key]
        if not subset:
            continue
        total   = len(subset)
        passed  = sum(1 for t in subset if t["status"] == "PASSED")
        failed  = sum(1 for t in subset if t["status"] == "FAILED")
        skipped = sum(1 for t in subset if t["status"] == "SKIPPED")
        rate    = f"{(passed / total * 100):.1f}%" if total else "N/A"
        ws.append([label, total, passed, failed, skipped, rate])
        fill = GREEN if failed == 0 else RED
        for col in range(1, 7):
            _fill_cell(ws.cell(ws.max_row, col), fill)

    _autowidth(ws)


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--vitest",  default="reports/vitest-results.json")
    parser.add_argument("--pytest",  default="reports/pytest-results.xml")
    parser.add_argument("--out",     default=None)
    parser.add_argument("--duration", type=float, default=0.0)
    args = parser.parse_args()

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ts_file   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    out_path = args.out or f"reports/test-report-{ts_file}.xlsx"
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)

    print(f"[report] Parsing Vitest: {args.vitest}")
    vitest_tests = parse_vitest(args.vitest)
    print(f"[report] Parsing pytest: {args.pytest}")
    pytest_tests = parse_pytest(args.pytest)

    all_tests = vitest_tests + pytest_tests
    print(f"[report] Total tests: {len(all_tests)} "
          f"({sum(1 for t in all_tests if t['status'] == 'PASSED')} passed, "
          f"{sum(1 for t in all_tests if t['status'] == 'FAILED')} failed)")

    wb = openpyxl.Workbook()
    build_summary(wb, all_tests, args.duration, timestamp)
    build_details(wb, all_tests)
    build_failures(wb, all_tests)
    build_by_layer(wb, all_tests)

    wb.save(out_path)
    print(f"[report] Saved -> {out_path}")
    return 1 if any(t["status"] == "FAILED" for t in all_tests) else 0


if __name__ == "__main__":
    sys.exit(main())
